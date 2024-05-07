import openpyxl

def es_palabra_clave(palabra, transiciones):
    estado_actual = 'q0'
    for caracter in palabra:
        if (estado_actual, caracter) in list(transiciones.keys()):
            estado_actual = transiciones[(estado_actual, caracter)]
        else:
            return False
    return estado_actual in estados_finales

# Función para procesar un archivo de texto línea por línea
def identificar_palabras_clave_en_texto(archivo, transiciones):
    historia = []  # Almacenar la historia del proceso
    contador_palabras_clave = {}  # Contador de palabras clave y sus ubicaciones

    with open(archivo, 'r') as archivo_texto:
        for num_linea, linea in enumerate(archivo_texto, start=1):
            palabras = linea.split()
            for num_palabra, palabra in enumerate(palabras, start=1):
                estado_final = es_palabra_clave(palabra, transiciones)
                if estado_final:
                    palabra_clave = estados_finales.get(estado_final)
                    if palabra_clave:
                        if palabra_clave not in contador_palabras_clave:
                            contador_palabras_clave[palabra_clave] = []

                        contador_palabras_clave[palabra_clave].append((num_linea, num_palabra))

                # Registra la historia del proceso
                historia.append(f"Caracter leído: '{palabra}'\tEstado actual: {estado_final}")
                print(historia)

    # Guarda la historia del proceso en un archivo
    with open('historia_del_proceso.txt', 'w') as historia_archivo:
        historia_archivo.write('\n'.join(historia))

    # Guarda el contador de palabras clave en un archivo
    with open('contador_palabras_clave.txt', 'w') as contador_archivo:
        for palabra, ubicaciones in contador_palabras_clave.items():
            contador_archivo.write(f"{palabra}: {len(ubicaciones)} encontradas\n")
            for ubicacion in ubicaciones:
                contador_archivo.write(f"- Encontrada en línea {ubicacion[0]}, palabra {ubicacion[1]}\n")

    print("Contador de palabras clave guardado en 'contador_palabras_clave.txt'")


def leer_transiciones(archivo_xlsx):
    transiciones = {}

    # Abre el archivo Excel
    wb = openpyxl.load_workbook(archivo_xlsx)
    ws = wb.active

    # Obtén los encabezados de la primera fila
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        estado_inicial = row[0]
        for index, estado_destino in enumerate(row[1:-1], start=1):
            entrada = headers[index]
            transiciones[(estado_inicial, entrada)] = estado_destino

    return transiciones

# Estados finales y sus palabras clave correspondientes
estados_finales = {
    'q5,q27,q0': 'if',
    'q38,q0': 'do',
    'q47,q0,q6': 'int',
    'q56,q0,q7': 'for',
    'q67,q0': 'enum',
    'q14,q69,q0': 'else',
    'q70,q0': 'goto',
    'q73,q0': 'auto',
    'q13,q85,q0': 'long',
    'q138,q0,q86': 'void',
    'q88,q14,q0': 'case',
    'q91,q0,q7': 'char',
    'q0,q96': 'union',
    'q97,q0': 'break',
    'q0,q102,q6': 'short',
    'q105,q0,q6': 'float',
    'q14,q0,q106': 'while',
    'q108,q0,q23,q6': 'const',
    'q112,q0': 'extern',
    'q138,q0,q133': 'signed',
    'q5,q0,q115': 'sizeof',
    'q1,q0,q116': 'static',
    'q0,q117,q6': 'struct',
    'q37,q0,q118': 'switch',
    'q0,q120': 'return',
    'q125,q14,q0': 'double',
    'q65,q5,q0,q129': 'typedef',
    'q132,q0,q6': 'default',
    'q138,q0,q133,q134': 'unsigned',
    'q0,q7,q135': 'register',
    'q14,q0,q136': 'volatile',
    'q137,q14,q0': 'continue'
}

archivo_xlsx = "buscador.xlsx"  # Reemplaza "tu_archivo.xlsx" por el nombre/path de tu archivo
transiciones = leer_transiciones(archivo_xlsx)

# Llamar a la función para identificar palabras clave en un archivo de texto
identificar_palabras_clave_en_texto('jambas.txt', transiciones)

