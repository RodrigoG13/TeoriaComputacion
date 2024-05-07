import openpyxl

def es_palabra_clave(palabra, transiciones):
    while palabra:
        estado_actual = 'q0'
        historia_temp = []
        for caracter in palabra:
            # Procesa el carácter con las transiciones del DFA
            if (estado_actual, caracter) in transiciones:
                estado_siguiente = transiciones[(estado_actual, caracter)]
                historia_temp.append(f"Carácter leído: '{caracter}'\tEstado actual: {estado_actual}\tEstado siguiente: {estado_siguiente}")
                estado_actual = estado_siguiente
            else:
                break  # Si no hay transición válida, salimos del bucle
        if estado_actual in estados_finales:
            return True, historia_temp
        palabra = palabra[1:]  # Elimina el primer carácter y vuelve a intentarlo
    return False, []


def identificar_palabras_clave_en_texto(archivo, transiciones):
    historia = []  # Almacenar la historia del proceso
    palabras_encontradas = {}  # Almacenar palabras reservadas y sus posiciones
    detalles_palabras = []  # Almacenar detalles de las palabras encontradas

    with open(archivo, 'r') as archivo_texto:
        for num_linea, linea in enumerate(archivo_texto, start=1):
            palabras = linea.split()
            for num_palabra, palabra in enumerate(palabras, start=1):
                es_clave, historia_temp = es_palabra_clave(palabra, transiciones)
                historia.extend(historia_temp)
                if es_clave:
                    detalles = f"Palabra clave ANSI C encontrada: {palabra} (línea {num_linea}, palabra {num_palabra})"
                    print(detalles)
                    detalles_palabras.append(detalles)
                    palabras_encontradas[palabra] = palabras_encontradas.get(palabra, 0) + 1

    # Guarda la historia del proceso en un archivo
    with open('historia_del_proceso.txt', 'w') as historia_archivo:
        historia_archivo.write('\n'.join(historia))

    # Guarda detalles de palabras clave encontradas
    with open('palabras_encontradas.txt', 'w') as archivo_palabras:
        for detalle in detalles_palabras:
            archivo_palabras.write(detalle + '\n')

    # Imprime el resumen de palabras reservadas encontradas
    print("\nResumen de palabras reservadas encontradas:")
    for palabra, conteo in palabras_encontradas.items():
        print(f"{palabra}: {conteo} veces")

def leer_transiciones(archivo_xlsx):
    transiciones = {}

    # Abre el archivo Excel
    wb = openpyxl.load_workbook(archivo_xlsx)
    ws = wb.active

    # Obtén los encabezados de la primera fila
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        estado_inicial = row[0]
        for index, estado_destino in enumerate(row[1:-1], start=1):
            entrada = headers[index]
            transiciones[(estado_inicial, entrada)] = estado_destino

    return transiciones

# Estados finales y sus palabras clave correspondientes
estados_finales = {
    'q5,q27,q0': 'if',
    'q38,q0': 'do',
    'q47,q0,q6': 'int',
    'q56,q0,q7': 'for',
    'q67,q0': 'enum',
    'q14,q69,q0': 'else',
    'q70,q0': 'goto',
    'q73,q0': 'auto',
    'q13,q85,q0': 'long',
    'q138,q0,q86': 'void',
    'q88,q14,q0': 'case',
    'q91,q0,q7': 'char',
    'q0,q96': 'union',
    'q97,q0': 'break',
    'q0,q102,q6': 'short',
    'q105,q0,q6': 'float',
    'q14,q0,q106': 'while',
    'q108,q0,q23,q6': 'const',
    'q112,q0': 'extern',
    'q138,q0,q133': 'signed',
    'q5,q0,q115': 'sizeof',
    'q1,q0,q116': 'static',
    'q0,q117,q6': 'struct',
    'q37,q0,q118': 'switch',
    'q0,q120': 'return',
    'q125,q14,q0': 'double',
    'q65,q5,q0,q129': 'typedef',
    'q132,q0,q6': 'default',
    'q138,q0,q133,q134': 'unsigned',
    'q0,q7,q135': 'register',
    'q14,q0,q136': 'volatile',
    'q137,q14,q0': 'continue'
}

archivo_xlsx = "buscador.xlsx"
transiciones = leer_transiciones(archivo_xlsx)
identificar_palabras_clave_en_texto('jambas.txt', transiciones)

