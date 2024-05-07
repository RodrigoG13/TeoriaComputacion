import openpyxl

def es_palabra_clave(palabra, transiciones):
    longitud = 0
    historia_temp = []
    estado_actual = 'q0'
    print(palabra)
    input()
    for idx, caracter in enumerate(palabra):
        if caracter.isalpha():
            if (estado_actual, caracter) in transiciones:
                estado_siguiente = transiciones[(estado_actual, caracter)]
                historia_temp.append(f"Carácter leído: '{caracter}'\tEstado actual: {estado_actual}\tEstado siguiente: {estado_siguiente}")
                estado_actual = estado_siguiente
                longitud += 1
            else:
                break

    if estado_actual in estados_finales and estados_finales[estado_actual] == palabra[:longitud]:
        return True, historia_temp, longitud
    return False, historia_temp, 0

def leer_transiciones(archivo_xlsx):
    transiciones = {}
    wb = openpyxl.load_workbook(archivo_xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        estado_inicial = row[0]
        for index, estado_destino in enumerate(row[1:-1], start=1):
            entrada = headers[index]
            transiciones[(estado_inicial, entrada)] = estado_destino

    return transiciones

def identificar_palabras_clave_en_texto(archivo, transiciones):
    historia = []
    palabras_encontradas = {}
    detalles_palabras = []

    with open(archivo, 'r') as archivo_texto:
        for num_linea, linea in enumerate(archivo_texto, start=1):
            idx_caracter = 0
            while idx_caracter < len(linea):
                if linea[idx_caracter].isalpha():
                    palabra_clave, historia_temp, longitud = detectar_palabra_clave(linea[idx_caracter:], transiciones)
                    historia.extend(historia_temp)
                    if palabra_clave:
                        detalles = f"Palabra clave ANSI C encontrada: {palabra_clave} (línea {num_linea}, posición {idx_caracter})"
                        print(detalles)
                        detalles_palabras.append(detalles)
                        palabras_encontradas[palabra_clave] = palabras_encontradas.get(palabra_clave, 0) + 1
                        idx_caracter += longitud - 1
                idx_caracter += 1

    with open('historia_del_proceso.txt', 'w') as historia_archivo:
        historia_archivo.write('\n'.join(historia))

    with open('palabras_encontradas.txt', 'w') as archivo_palabras:
        for detalle in detalles_palabras:
            archivo_palabras.write(detalle + '\n')

    print("\nResumen de palabras reservadas encontradas:")
    for palabra, conteo in palabras_encontradas.items():
        print(f"{palabra}: {conteo} veces")

def detectar_palabra_clave(subcadena, transiciones):
    estados_actuales = {'q0'}
    historia_temp = []
    idx = 0
    while idx < len(subcadena) and subcadena[idx].isalpha():
        caracter = subcadena[idx]
        nuevos_estados = set()

        for estado_actual in estados_actuales:
            if (estado_actual, caracter) in transiciones:
                estado_siguiente = transiciones[(estado_actual, caracter)]
                if estado_siguiente:  # Ignoramos transiciones a estado None
                    nuevos_estados.add(estado_siguiente)

        if not nuevos_estados:
            break

        detalles_transicion = f"Carácter leído: '{caracter}'\tEstados actuales: {', '.join(estados_actuales)}\tEstados siguientes: {', '.join(nuevos_estados)}"
        historia_temp.append(detalles_transicion)

        estados_actuales = nuevos_estados
        idx += 1

    estado_final_valido = [estado for estado in estados_actuales if estado in estados_finales]
    if estado_final_valido:
        return estados_finales[estado_final_valido[0]], historia_temp, idx
    return None, historia_temp, idx

# Estados finales y sus palabras clave correspondientes
estados_finales = {
    'q5,q27,q0': 'if',
    'q38,q0': 'do',
    'q47,q0,q6': 'int',
    'q56,q0,q7': 'for',
    'q67,q0': 'enum',
    'q14,q69,q0': 'else',
    'q70,q0': 'goto',
    'q73,q0': 'auto',
    'q13,q85,q0': 'long',
    'q138,q0,q86': 'void',
    'q88,q14,q0': 'case',
    'q91,q0,q7': 'char',
    'q0,q96': 'union',
    'q97,q0': 'break',
    'q0,q102,q6': 'short',
    'q105,q0,q6': 'float',
    'q14,q0,q106': 'while',
    'q108,q0,q23,q6': 'const',
    'q112,q0': 'extern',
    'q138,q0,q133': 'signed',
    'q5,q0,q115': 'sizeof',
    'q1,q0,q116': 'static',
    'q0,q117,q6': 'struct',
    'q37,q0,q118': 'switch',
    'q0,q120': 'return',
    'q125,q14,q0': 'double',
    'q65,q5,q0,q129': 'typedef',
    'q132,q0,q6': 'default',
    'q138,q0,q133,q134': 'unsigned',
    'q0,q7,q135': 'register',
    'q14,q0,q136': 'volatile',
    'q137,q14,q0': 'continue'
}

archivo_xlsx = "buscador.xlsx"
transiciones = leer_transiciones(archivo_xlsx)
identificar_palabras_clave_en_texto('jambas.txt', transiciones)

